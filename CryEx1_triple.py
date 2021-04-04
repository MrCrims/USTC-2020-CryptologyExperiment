import matplotlib.pyplot as plt
from matplotlib.font_manager import FontProperties
import  matplotlib
import openpyxl
matplotlib.rcParams['font.sans-serif'] = ['SimHei']
matplotlib.rcParams['axes.unicode_minus'] = False
#
def read_message(txt_name):#从txt文档中读取明文,将所有明文并在一起，不分段
    temp=''
    with open('D:\CryExperiments\Ex1_Ciph.txt','r') as f:
        for i in f:
            temp += i.lower()
            temp = temp.replace(',', '')
            temp = temp.replace(' ', '')
            temp = temp.replace('!', '')
            temp = temp.replace('?', '')
            temp = temp.replace('.', '')
            temp = temp.replace('-', '')
            temp = temp.replace('"', '')
            temp = temp.replace(':', '')
            temp = temp.replace('(', '')
            temp = temp.replace(')', '')
            temp = temp.replace('\'', '')
            temp = temp.replace(';', '')
            temp = temp.replace('`', '')
            temp = temp.strip('\n')
    f.close()
    return temp
file_name = input('Please put in the file name:')
passage = read_message(file_name)
#
def count_triple_char(passage):#统计三字符
    dic={}
    for i in range(0,26):
        for j in range(0,26):
            for k in range(0,26):
                key=chr(i+97)+chr(j+97)+chr(k+97)
                dic[key]=0
    for i in range(0,len(passage)-2):
        key=passage[i]+passage[i+1]+passage[i+2]
        dic[key]+=1
    return dic
dic_triple=count_triple_char(passage)
#
def triple_char_xlsx(dic_triple):#将统计数据写入到excel中,之后交由matlab来画图
    f = openpyxl.Workbook()
    worksheet=f.create_sheet()
    for i in range(0,26):
        for j in range(0,26):
            for k in range(0,26):
                worksheet.cell(1+k+j*26+i*26*26,1,chr(i+97))
                worksheet.cell(1+k+j*26+i*26*26,2,chr(j+97))
                worksheet.cell(1+k+j*26+i*26*26,3,chr(k+97))
                worksheet.cell(1+k+j*26+i*26*26,4,dic_triple[chr(i+97)+chr(j+97)+chr(k+97)])
    f.save('D:\CryExperiments\Ex1_Triple.xlsx')
triple_char_xlsx(dic_triple)